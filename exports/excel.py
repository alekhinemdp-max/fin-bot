"""Генератор профессионального Excel-отчёта"""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart


# Цветовая палитра
C_HEADER_BG    = "1A1A2E"   # тёмно-синий
C_HEADER_FG    = "FFFFFF"   # белый
C_INCOME_BG    = "D4EDDA"   # светло-зелёный
C_EXPENSE_BG   = "F8D7DA"   # светло-красный
C_TOTAL_BG     = "FFF3CD"   # жёлтый
C_SUMMARY_BG   = "E8F4FD"   # светло-синий
C_STRIPE       = "F8F9FA"   # серая полоска
C_BORDER       = "DEE2E6"   # цвет границ
C_INCOME_TEXT  = "155724"
C_EXPENSE_TEXT = "721C24"
C_ACCENT       = "0D6EFD"   # синий акцент


def make_border(color=C_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def make_header_font(size=11, bold=True):
    return Font(name="Arial", size=size, bold=bold, color=C_HEADER_FG)


def make_body_font(size=10, bold=False, color="000000"):
    return Font(name="Arial", size=size, bold=bold, color=color)


def apply_header(cell, text, size=11):
    cell.value = text
    cell.font = make_header_font(size)
    cell.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = make_border()


def generate_excel_report(rows: list, label: str, date_from: str, date_to: str) -> bytes:
    wb = Workbook()

    # ─── Лист 1: Все транзакции ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Транзакции"

    # Заголовок отчёта
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"💰 Финансовый отчёт — {label}"
    title_cell.font = Font(name="Arial", size=14, bold=True, color=C_HEADER_FG)
    title_cell.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Период: {date_from}  →  {date_to}   |   Сгенерировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    sub_cell.font = Font(name="Arial", size=9, color="666666")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Пустая строка
    ws.row_dimensions[3].height = 6

    # Заголовки колонок
    headers = ["#", "Дата", "Время", "Тип", "Сумма", "Валюта", "Категория", "Описание"]
    for col, h in enumerate(headers, 1):
        apply_header(ws.cell(row=4, column=col), h, size=10)
    ws.row_dimensions[4].height = 22

    # Данные
    income_total = {}
    expense_total = {}
    
    for i, row in enumerate(rows):
        tx_id, tx_type, amount, currency, category, desc, created_at = row
        r = i + 5
        
        # Парсим дату
        try:
            dt = datetime.strptime(created_at[:19], "%Y-%m-%d %H:%M:%S")
            date_str = dt.strftime("%d.%m.%Y")
            time_str = dt.strftime("%H:%M")
        except:
            date_str = created_at[:10] if created_at else ""
            time_str = created_at[11:16] if created_at and len(created_at) > 10 else ""
        
        is_income = tx_type == "income"
        row_bg = C_INCOME_BG if is_income else C_EXPENSE_BG
        
        # Чётные строки — чуть темнее (зебра внутри цвета)
        if i % 2 == 1:
            row_bg = "C3E6CB" if is_income else "F5C6CB"
        
        cells_data = [
            (i + 1, "center"),
            (date_str, "center"),
            (time_str, "center"),
            ("📥 Доход" if is_income else "📤 Расход", "center"),
            (amount, "right"),
            (currency, "center"),
            (category or "Прочее", "left"),
            (desc or "", "left"),
        ]
        
        for col, (val, align) in enumerate(cells_data, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = make_body_font(
                bold=(col == 5),
                color=C_INCOME_TEXT if is_income else C_EXPENSE_TEXT
            )
            c.fill = PatternFill("solid", fgColor=row_bg)
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.border = make_border()
            
            if col == 5:  # Сумма
                c.number_format = '#,##0.00'
        
        ws.row_dimensions[r].height = 18
        
        # Считаем итоги
        if is_income:
            income_total[currency] = income_total.get(currency, 0) + amount
        else:
            expense_total[currency] = expense_total.get(currency, 0) + amount

    # Ширина колонок
    col_widths = [5, 12, 8, 14, 14, 9, 22, 40]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A5"

    # ─── Лист 2: Сводка ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Сводка")

    # Заголовок
    ws2.merge_cells("A1:D1")
    ws2["A1"].value = f"Финансовая сводка — {label}"
    ws2["A1"].font = Font(name="Arial", size=13, bold=True, color=C_HEADER_FG)
    ws2["A1"].fill = PatternFill("solid", fgColor=C_HEADER_BG)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    row_cur = 3

    # Доходы по валютам
    apply_header(ws2.cell(row=row_cur, column=1), "📥  ДОХОДЫ", 10)
    apply_header(ws2.cell(row=row_cur, column=2), "Сумма", 10)
    row_cur += 1
    
    if income_total:
        for curr, total in sorted(income_total.items()):
            c1 = ws2.cell(row=row_cur, column=1, value=curr)
            c2 = ws2.cell(row=row_cur, column=2, value=total)
            c1.fill = PatternFill("solid", fgColor=C_INCOME_BG)
            c2.fill = PatternFill("solid", fgColor=C_INCOME_BG)
            c1.font = make_body_font(color=C_INCOME_TEXT, bold=True)
            c2.font = make_body_font(color=C_INCOME_TEXT, bold=True)
            c2.number_format = '#,##0.00'
            c1.border = c2.border = make_border()
            c1.alignment = Alignment(horizontal="center")
            c2.alignment = Alignment(horizontal="right")
            row_cur += 1
    else:
        ws2.cell(row=row_cur, column=1, value="— нет данных —")
        row_cur += 1

    row_cur += 1  # отступ

    # Расходы по валютам
    apply_header(ws2.cell(row=row_cur, column=1), "📤  РАСХОДЫ", 10)
    apply_header(ws2.cell(row=row_cur, column=2), "Сумма", 10)
    row_cur += 1
    
    if expense_total:
        for curr, total in sorted(expense_total.items()):
            c1 = ws2.cell(row=row_cur, column=1, value=curr)
            c2 = ws2.cell(row=row_cur, column=2, value=total)
            c1.fill = PatternFill("solid", fgColor=C_EXPENSE_BG)
            c2.fill = PatternFill("solid", fgColor=C_EXPENSE_BG)
            c1.font = make_body_font(color=C_EXPENSE_TEXT, bold=True)
            c2.font = make_body_font(color=C_EXPENSE_TEXT, bold=True)
            c2.number_format = '#,##0.00'
            c1.border = c2.border = make_border()
            c1.alignment = Alignment(horizontal="center")
            c2.alignment = Alignment(horizontal="right")
            row_cur += 1
    else:
        ws2.cell(row=row_cur, column=1, value="— нет данных —")
        row_cur += 1

    row_cur += 1  # отступ

    # Баланс ARS
    inc_ars = income_total.get("ARS", 0)
    exp_ars = expense_total.get("ARS", 0)
    balance = inc_ars - exp_ars
    
    balance_label = ws2.cell(row=row_cur, column=1, value="💼  БАЛАНС (ARS)")
    balance_val   = ws2.cell(row=row_cur, column=2, value=balance)
    
    balance_color = "D4EDDA" if balance >= 0 else "F8D7DA"
    balance_text_color = C_INCOME_TEXT if balance >= 0 else C_EXPENSE_TEXT
    
    for c in [balance_label, balance_val]:
        c.fill = PatternFill("solid", fgColor=balance_color)
        c.font = Font(name="Arial", size=11, bold=True, color=balance_text_color)
        c.border = make_border()
    balance_val.number_format = '+#,##0.00;-#,##0.00'
    balance_label.alignment = Alignment(horizontal="center", vertical="center")
    balance_val.alignment = Alignment(horizontal="right", vertical="center")
    ws2.row_dimensions[row_cur].height = 22
    row_cur += 2

    # Топ категорий расходов
    if rows:
        ws2.cell(row=row_cur, column=1).value = "🏷  ТОП КАТЕГОРИИ РАСХОДОВ"
        ws2.cell(row=row_cur, column=2).value = "Сумма ARS"
        for col in [1, 2]:
            apply_header(ws2.cell(row=row_cur, column=col), ws2.cell(row=row_cur, column=col).value, 10)
        row_cur += 1
        
        cats = {}
        for row in rows:
            tx_id, tx_type, amount, currency, category, desc, created_at = row
            if tx_type == "expense" and currency == "ARS":
                cats[category or "Прочее"] = cats.get(category or "Прочее", 0) + amount
        
        for cat, total in sorted(cats.items(), key=lambda x: x[1], reverse=True)[:10]:
            c1 = ws2.cell(row=row_cur, column=1, value=cat)
            c2 = ws2.cell(row=row_cur, column=2, value=total)
            stripe = C_STRIPE if row_cur % 2 == 0 else "FFFFFF"
            c1.fill = c2.fill = PatternFill("solid", fgColor=stripe)
            c1.font = c2.font = make_body_font()
            c1.border = c2.border = make_border()
            c2.number_format = '#,##0.00'
            c2.alignment = Alignment(horizontal="right")
            row_cur += 1

    # Ширина колонок сводки
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 16

    # ─── Сохраняем ───────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
